# generator.py
import json
import re
from math import floor
from typing import Optional, List
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ------------------ 工具：向合并单元格写值 ------------------
def write_merged(ws, range_str: str, value):
    """向一个合并单元格区域写入值（例如 'E1:F1'），先解合并再恢复。"""
    try:
        ws.unmerge_cells(range_str)
    except Exception:
        pass
    top_left = range_str.split(":")[0]
    ws[top_left].value = value
    ws.merge_cells(range_str)

# ------------------ 文本清洗 ------------------
def strip_brackets(text: str) -> str:
    if not text:
        return ""
    # 去掉各种括号外壳： 【】[]（）()
    return re.sub(r"[【】\[\]（）()]", "", str(text)).strip()

def after_colon(text: str) -> str:
    """拿到 'xxx：内容' 里的 '内容'；如果没有顿号，就直接返回原文"""
    if not text:
        return ""
    s = str(text)
    return s.split("：", 1)[1].strip() if "：" in s else s.strip()

# ------------------ 技能表检索 ------------------
def find_skill_row(sheet, skill_id: str) -> Optional[int]:
    """在技能sheet的B列中找到等于 skill_id 的行号，找不到返回None"""
    target = str(skill_id).strip()
    for r in range(1, sheet.max_row + 1):
        val = sheet.cell(row=r, column=2).value  # B列
        if val is None:
            continue
        if str(val).strip() == target:
            return r
    return None

# ------------------ 属性区行号 ------------------
FIELDS = [
    ("hp", 4),
    ("me", 5),
    ("str", 6),
    ("agi", 7),
    ("wil", 8),
    ("obs", 9),
    ("wis", 10),
    ("cha", 11),
]

# ------------------ 写入单个角色 ------------------
def write_character(ws, right_col, data, skills_wb):
    """
    right_col: 当前角色右列列号（E/F -> F列=6；G/H -> 8；以此类推）
    skills_wb: 技能总表工作簿
    """
    left_col = right_col - 1
    L = get_column_letter(left_col)
    R = get_column_letter(right_col)

    # ---- 名称 / 职业 ----
    write_merged(ws, f"{L}1:{R}1", data.get("name", ""))
    write_merged(ws, f"{L}2:{R}2", data.get("seqName", ""))
    # 第3行（玩家）留空

    # ---- 属性 ----
    for key, row in FIELDS:
        v = data.get(key, 0)
        left_v = v if key in ("hp", "me") else floor(v / 5)
        ws.cell(row=row, column=left_col, value=left_v)
        ws.cell(row=row, column=right_col, value=v)

    # ---- 生活技能（15、16行）----
    pathway_id = str(data.get("pathwayId", "")).lstrip("0")
    if pathway_id in skills_wb.sheetnames:
        sh = skills_wb[pathway_id]
        a3 = sh["A3"].value or ""
        b3 = (sh["B3"].value or "").strip()

        # 取A3中“生活技能：”后的内容，并去括号
        skill_brief = strip_brackets(after_colon(a3))
        write_merged(ws, f"{L}15:{R}15", f"生活技能：{skill_brief}")
        write_merged(ws, f"{L}16:{R}16", b3)
    else:
        write_merged(ws, f"{L}15:{R}15", "生活技能：")
        write_merged(ws, f"{L}16:{R}16", f"(未找到技能Sheet：{pathway_id})")
        return  # 没有sheet就不继续技能循环了

    # ---- 技能循环（从17行开始，每个技能占6行：17~22，然后23~28...）----
    skills: List[str] = list(data.get("skillIds", []))
    out_row = 17  # 当前块的第一行（技能名称行）

    for sid in skills:
        if not sid:
            continue

        # 目标sheet = pathwayId(去前导0)
        sh = skills_wb[pathway_id]
        base = find_skill_row(sh, sid)

        if base is None:
            # 没找到该技能，写个提示避免空白
            write_merged(ws, f"{L}{out_row}:{R}{out_row}", f"技能未找到：{sid}")
            # 留空后续几行
            for rr in range(out_row + 1, out_row + 6):
                write_merged(ws, f"{L}{rr}:{R}{rr}", "")
            out_row += 6
            continue

        # 映射：基于“技能序号所在行 base”，按你的偏移读取B列
        # 17：技能名称 = base+1
        name_raw = sh.cell(row=base + 1, column=2).value or ""
        skill_name = strip_brackets(after_colon(name_raw))
        write_merged(ws, f"{L}{out_row}:{R}{out_row}", skill_name)

        # 18：效果 = base+5
        effect_raw = sh.cell(row=base + 5, column=2).value or ""
        effect = after_colon(effect_raw)
        write_merged(ws, f"{L}{out_row+1}:{R}{out_row+1}", effect)

        # 19：ME消耗 = base+7
        me_raw = sh.cell(row=base + 7, column=2).value or ""
        me_cost = after_colon(me_raw)
        write_merged(ws, f"{L}{out_row+2}:{R}{out_row+2}", me_cost)

        # 20：技能范围 = base+8
        range_raw = sh.cell(row=base + 8, column=2).value or ""
        range_txt = after_colon(range_raw)
        write_merged(ws, f"{L}{out_row+3}:{R}{out_row+3}", range_txt)

        # 21：持续时间 = base+10
        dur_raw = sh.cell(row=base + 10, column=2).value or ""
        duration = after_colon(dur_raw)
        write_merged(ws, f"{L}{out_row+4}:{R}{out_row+4}", duration)

        # 22：判定 = base+6
        judge_raw = sh.cell(row=base + 6, column=2).value or ""
        judge = after_colon(judge_raw)
        write_merged(ws, f"{L}{out_row+5}:{R}{out_row+5}", judge)

        # 下一组技能块
        out_row += 6

# ------------------ 入口 ------------------
def generate_excel(template_path, json_files, skills_path, output_path="成品输出.xlsx"):
    wb = load_workbook(template_path)
    ws = wb.active
    skills_wb = load_workbook(skills_path, data_only=True)

    # === 读取所有角色数据并排序 ===
    all_data = []
    for file_path in json_files:
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            all_data.append((data.get("agi", 0), file_path, data))

    # 按敏捷（agi）降序排序
    all_data.sort(key=lambda x: x[0], reverse=True)

    # === 列位：E/F 起步，每个角色占两列 ===
    start_right_col = 6
    step = 2

    for i, (_, file_path, data) in enumerate(all_data):
        print(f"写入角色：{data.get('name','未知')}（敏捷={data.get('agi',0)}）")
        write_character(ws, start_right_col + i * step, data, skills_wb)

    wb.save(output_path)
    print(f"✅ 成品生成完成: {output_path}")
