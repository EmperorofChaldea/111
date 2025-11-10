from openpyxl import load_workbook

path = "成品.xlsx"     # 你的模板文件
wb = load_workbook(path)
ws = wb.active

# 目标合并单元格范围
target_range = "E1:F1"

# 先解合并
ws.unmerge_cells(target_range)

# 写入左上角单元格
ws["E1"].value = "史蒂芬"

# 再重新合并
ws.merge_cells(target_range)

wb.save("成品测试输出.xlsx")
print("✅ 已成功写入 E1:F1 = '史蒂芬'")
